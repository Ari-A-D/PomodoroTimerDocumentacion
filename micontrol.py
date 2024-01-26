import os
import time
import pickle
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import PySimpleGUI as sg
import winsound

# Configuración de tiempos en milisegundos
tiempo_documentar = 1500000 #25 = 1500000 minutos
horas_cumplir = 300  # 300 horas

class Timer:
    def __init__(self, hours):
        self.start_time = None
        self.remaining_time = hours * 3600  # Convertir horas a segundos

    def start(self):
        self.start_time = time.time()

    def pause(self):
        if self.start_time is not None:
            elapsed_time = time.time() - self.start_time
            self.remaining_time -= elapsed_time
            self.start_time = None

    def stop(self):
        self.pause()
        self.save_state()

    def save_state(self):
        with open('timer_state.pkl', 'wb') as f:
            pickle.dump(self.remaining_time, f)

def load_state():
    try:
        with open('timer_state.pkl', 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return None

remaining_time = load_state()
if remaining_time is not None:
    timer = Timer(remaining_time / 3600)  # Convertir segundos a horas
else:
    timer = Timer(horas_cumplir)  # Comenzar con 300 horas

def update_label():
    if timer.start_time:
        elapsed_time = time.time() - timer.start_time
        timer.remaining_time -= elapsed_time
        timer.start_time = time.time()
    remaining_seconds = int(timer.remaining_time)
    hours, remaining_seconds = divmod(remaining_seconds, 3600)
    minutes, seconds = divmod(remaining_seconds, 60)
    label.config(text=f'{hours}:{minutes:02}:{seconds:02}')
    root.after(1000, update_label)  # Llamar a update_label() cada 1000 milisegundos

def reset_button_styles():
    start_button.config(bg="green", fg="white")
    start_button_border.config(highlightbackground="black", highlightthickness=0)
    pause_button.config(bg="orange", fg="white")
    pause_button_border.config(highlightbackground="black", highlightthickness=0)
    stop_button.config(bg="red", fg="white")
    stop_button_border.config(highlightbackground="black", highlightthickness=0)

def show_popup():
    global popup 
    popup = tk.Toplevel()
    popup.title("¿QUÉ ESTÁS HACIENDO?")
    popup.configure(bg="black")

    label = tk.Label(popup, text="Documenta:", bg="black", fg="lime", width=20, height=5)
    label.pack()

    entry = tk.Entry(popup, width=50)
    entry.pack()

    def play_sound():
        winsound.PlaySound("path/to/sound.wav", winsound.SND_ASYNC)

    play_sound()

    def close_popup():
        popup.destroy()

    save_button = tk.Button(popup, text="Guardar", command=lambda: [save_text(entry), close_popup()], bg="black", fg="lime", width=20, height=2)
    save_button.pack()

def save_text(entry):
    text = entry.get()

    layout = [
        [sg.VPush()],
        [sg.Push(), sg.Text('Selecciona uno:'), sg.Push()],
        [sg.Push(), sg.Column([[sg.Button('Remoto', size=(10, 2)), sg.Button('Presencial', size=(10, 2))]], justification='center', element_justification='center'), sg.Push()],
        [sg.VPush()]
    ]

    window = sg.Window('Selecciona el modo de trabajo', layout, size=(300, 200))
    event, _ = window.read()
    window.close()

    if event == 'Remoto':
        mode = 'Remoto'
    elif event == 'Presencial':
        mode = 'Presencial'
    else:
        return

    current_date = time.strftime("%Y-%m-%d")
    folder_path = os.path.join("bitacora", current_date)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    file_path = os.path.join(folder_path, f"{current_date}.xlsx")

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Fecha-Hora"
        sheet["B1"] = "Modo"
        sheet["C1"] = "Descripcion"

    row = [time.strftime("%Y-%m-%d %H:%M:%S"), mode, text]
    sheet.append(row)

    workbook.save(file_path)
    messagebox.showinfo("Guardado", "Ya está documentado")

timer_paused = False
timer_stopped = False

def schedule_popup():
    if timer_paused or timer_stopped:
        return

    show_popup()
    root.after(tiempo_documentar, schedule_popup)   # Cada 10 segundos

def start_timer():
    global timer_paused, timer_stopped
    reset_button_styles()
    start_button.config(bg="green", fg="white")
    start_button_border.config(highlightbackground="green", highlightthickness=5)
    timer.start()
    update_label()

    timer_paused = False
    timer_stopped = False
    schedule_popup()

def pause_timer():
    global timer_paused
    reset_button_styles()
    pause_button.config(bg="orange", fg="white")
    pause_button_border.config(highlightbackground="orange", highlightthickness=5)
    timer.pause()
    update_label()

    if 'popup' in globals() and popup.winfo_exists():
        popup.destroy()

    timer_paused = True

def stop_timer():
    global timer_stopped
    reset_button_styles()
    stop_button.config(bg="red", fg="white")
    stop_button_border.config(highlightbackground="red", highlightthickness=5)
    timer.stop()
    update_label()

    if 'popup' in globals() and popup.winfo_exists():
        popup.destroy()

    timer_stopped = True

root = tk.Tk()
root.title("Temporizador")
root.configure(bg="black")

label = tk.Label(root, text="", font=('Helvetica', 48), bg="black", fg="white")
label.pack()

start_button_border = tk.Frame(root, bg="black", highlightbackground="black", highlightthickness=0)
start_button_border.pack(pady=5)
start_button = tk.Button(start_button_border, text="Iniciar", command=start_timer, bg="green", fg="white", relief="raised", padx=10, pady=5, bd=0, width=10, borderwidth=0, highlightthickness=0)
start_button.pack()

pause_button_border = tk.Frame(root, bg="black", highlightbackground="black", highlightthickness=0)
pause_button_border.pack(pady=5)
pause_button = tk.Button(pause_button_border, text="Pausar", command=pause_timer, bg="orange", fg="white", relief="raised", padx=10, pady=5, bd=0, width=10, borderwidth=0, highlightthickness=0)
pause_button.pack()

stop_button_border = tk.Frame(root, bg="black", highlightbackground="black", highlightthickness=0)
stop_button_border.pack(pady=5)
stop_button = tk.Button(stop_button_border, text="Detener", command=stop_timer, bg="red", fg="white", relief="raised", padx=10, pady=5, bd=0, width=10, borderwidth=0, highlightthickness=0)
stop_button.pack()

update_label()

root.mainloop()
